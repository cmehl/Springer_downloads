import os
import time

import openpyxl

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys


# creating folder where books will be stored
if not os.path.isdir("./books"):
	os.makedirs('./books')


# Opening workbook with book information
wb = openpyxl.load_workbook('Free+English+textbooks.xlsx')
sheet = wb.active


# Selenium browser
# Options modified to trigger download at click and save it in books folder
options = webdriver.ChromeOptions()
options.add_experimental_option('prefs', {
"download.default_directory": '/Users/mehlcedric/Documents/Personnel/Misc/Springer_downloads/books', #Change default directory for downloads
"download.prompt_for_download": False, #To auto download the file
"download.directory_upgrade": True,
"plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
})
#
browser = webdriver.Chrome(ChromeDriverManager().install(), options=options)

# URLs are in columns S
failed = 0
list_failed = []
for rowOfCellObjects in sheet['S2':f'S{sheet.max_row}']:
	for cellObj in rowOfCellObjects:   # loop on cells in the row

		# Reading excel data
		URL = cellObj.value
		row_nb = cellObj.row
		name = sheet.cell(row=row_nb, column=1).value
		print(f"Downloading {name}")

		# Open URL in browser
		browser.get(URL)

		# Download link
		linkElem = browser.find_element_by_partial_link_text('Download')
		try:
			linkElem.send_keys("\n")
		except:
			failed += 1
			list_failed.append(name)

# Quit browser (after 30s pausing to make sure all downloads are made)
time.sleep(30)
browser.quit()

# Outputting some information if some downloads failed
print(f"\nWARNING: {failed} downloads failed")
if failed>0:
	print("The missing PDF's are:")
	for name in list_failed:
		print(f"- {name}")













